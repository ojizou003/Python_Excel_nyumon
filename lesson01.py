import openpyxl
import glob

# test.xlsxというファイル名のExcelファイルを作成
# wb = openpyxl.Workbook()
# wb.active
# wb.save('test.xlsx')

# # Excelファイルを10枚作成
# for i in range(10):
#     file_name = 'test_file{}.xlsx'.format(i)
#     wb = openpyxl.Workbook()
#     wb.active
#     wb.save(file_name)

# ファイル一覧を取得する
# print(glob.glob('test_file*'))

# for文を使って見やすく
# for i, file_name in enumerate(glob.glob('test_file*')):
#     print(i,file_name)

# test.xlsxを読み込む
# wb = openpyxl.load_workbook('Excelfiles/test.xlsx')

# シート名の確認
# print(wb.sheetnames)

# # ファイルの保存
# wb.save('Excelfiles/test2.xlsx')

